import os
import time
import json
import uuid
import pathlib
from typing import Optional, Dict, Any, List
from urllib.parse import urlencode

import httpx
from dotenv import load_dotenv
from fastapi import FastAPI, Request, HTTPException, UploadFile, File, Form
from fastapi.responses import RedirectResponse, PlainTextResponse, HTMLResponse
from starlette.middleware.sessions import SessionMiddleware
from passlib.context import CryptContext

# Excel reading
from openpyxl import load_workbook

# ----------------------------
# Config / environment
# ----------------------------
load_dotenv()

HMRC_CLIENT_ID = os.getenv("HMRC_CLIENT_ID", "")
HMRC_CLIENT_SECRET = os.getenv("HMRC_CLIENT_SECRET", "")
HMRC_REDIRECT_URI = os.getenv("HMRC_REDIRECT_URI", "http://localhost:3000/oauth/hmrc/callback")
BASE_URL = os.getenv("BASE_URL", "https://test-api.service.hmrc.gov.uk")
SCOPE = "read:vat write:vat read:vat-returns"

# Session
SESSION_SECRET = os.getenv("SESSION_SECRET", "please_change_me_32chars")

# Data dir (use /data on Render if you mounted a disk)
DATA_DIR = pathlib.Path(os.getenv("DATA_DIR", "."))
DATA_DIR.mkdir(parents=True, exist_ok=True)

# Files
TOKEN_FILE = DATA_DIR / "tokens.json"
RECEIPTS_FILE = DATA_DIR / "receipts.json"
USERS_FILE = DATA_DIR / "users.json"   # NEW: user store (email + hash + role + profile)

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# ----------------------------
# Persistence helpers
# ----------------------------
def load_json(path: pathlib.Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text())
        except Exception:
            return default
    return default

def save_json(path: pathlib.Path, data: Any):
    path.write_text(json.dumps(data, indent=2))

def load_tokens():
    return load_json(TOKEN_FILE, None)

def save_tokens(tokens: dict):
    save_json(TOKEN_FILE, tokens)

def append_receipt(vrn: str, period_key: Optional[str], receipt: dict):
    data = load_json(RECEIPTS_FILE, [])
    data.append({"vrn": vrn, "periodKey": period_key, **receipt})
    save_json(RECEIPTS_FILE, data)

def load_users() -> List[Dict[str, Any]]:
    return load_json(USERS_FILE, [])

def save_users(users: List[Dict[str, Any]]):
    save_json(USERS_FILE, users)

def get_user_by_email(email: str) -> Optional[Dict[str, Any]]:
    email = (email or "").strip().lower()
    for u in load_users():
        if u.get("email", "").lower() == email:
            return u
    return None

def create_user(*, email: str, password: str, role: str, contact_name: str="", business_name: str="", phone: str="") -> Dict[str, Any]:
    if get_user_by_email(email):
        raise ValueError("An account with this email already exists.")
    if any(c in password for c in "<>,&\"'"):
        raise ValueError("Please do not use < > , & \" ' in passwords.")
    if len(password) < 8:
        raise ValueError("Password must be at least 8 characters.")
    hashed = pwd_context.hash(password)
    user = {
        "id": str(uuid.uuid4()),
        "email": email.strip().lower(),
        "hash": hashed,
        "role": role,                  # 'agent' | 'taxpayer'
        "contact_name": contact_name,
        "business_name": business_name,
        "phone": phone,
        "created_at": int(time.time()),
    }
    users = load_users()
    users.append(user)
    save_users(users)
    return user

def verify_password(password: str, user: Dict[str, Any]) -> bool:
    return pwd_context.verify(password, user.get("hash", ""))

# ----------------------------
# App + session
# ----------------------------
app = FastAPI()
app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)

STORE = {"tokens": load_tokens(), "state": None, "device_id": str(uuid.uuid4())}

# ----------------------------
# Helpers (HMRC + auth)
# ----------------------------
def hmrc_headers(request: Request) -> dict:
    """
    Minimal-but-valid Fraud Prevention headers for sandbox.
    Expand for production.
    """
    user_agent = request.headers.get("user-agent", "my-vat-filer/1.0")
    client_ip = request.client.host if request.client else "203.0.113.10"
    return {
        "Accept": "application/vnd.hmrc.1.0+json",
        "User-Agent": "my-vat-filer/1.0",
        "Gov-Client-Device-Id": STORE["device_id"],
        "Gov-Client-Public-IP": client_ip,
        "Gov-Client-Local-IPs": "192.168.1.10",
        "Gov-Client-Timezone": "UTC+00:00",
        "Gov-Client-User-IDs": "os=user123",
        "Gov-Vendor-Version": "my-vat-filer=1.0.0",
        "Gov-Client-User-Agent": user_agent,
    }

def auth_url(state: str) -> str:
    q = {
        "response_type": "code",
        "client_id": HMRC_CLIENT_ID,
        "redirect_uri": HMRC_REDIRECT_URI,
        "scope": SCOPE,
        "state": state,
    }
    return f"{BASE_URL}/oauth/authorize?{urlencode(q)}"

async def token_request(data: dict) -> dict:
    async with httpx.AsyncClient() as client:
        r = await client.post(
            f"{BASE_URL}/oauth/token",
            data=data,
            headers={"Content-Type": "application/x-www-form-urlencoded"},
        )
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)
    return r.json()

async def access_token() -> str:
    """
    Return a valid access token; auto-refresh it if it's near expiry.
    """
    t = STORE["tokens"]
    if not t:
        raise HTTPException(401, "Not connected to HMRC yet.")

    # Refresh 60 seconds before expiry
    obtained = t.get("obtained_at")
    expires_in = t.get("expires_in", 0)
    if obtained and (obtained + expires_in - 60) < time.time():
        async with httpx.AsyncClient() as client:
            r = await client.post(
                f"{BASE_URL}/oauth/token",
                data={
                    "grant_type": "refresh_token",
                    "client_id": HMRC_CLIENT_ID,
                    "client_secret": HMRC_CLIENT_SECRET,
                    "refresh_token": t["refresh_token"],
                },
                headers={"Content-Type": "application/x-www-form-urlencoded"},
            )
        if r.status_code >= 400:
            raise HTTPException(r.status_code, r.text)
        t = r.json()
        t["obtained_at"] = time.time()
        STORE["tokens"] = t
        save_tokens(t)
    return t["access_token"]

def current_user(request: Request) -> Optional[Dict[str, Any]]:
    user = request.session.get("user")
    if not user:
        return None
    # Refresh user details from store
    u = get_user_by_email(user.get("email", ""))
    return {"email": u["email"], "role": u["role"], "contact_name": u.get("contact_name"), "business_name": u.get("business_name")} if u else None

def require_login(request: Request, roles: Optional[List[str]] = None) -> Optional[RedirectResponse]:
    """
    Redirect to /login if not logged in.
    Optionally restrict roles: roles=['agent'] or ['taxpayer']
    """
    u = current_user(request)
    if not u:
        return RedirectResponse("/login", status_code=303)
    if roles and u["role"] not in roles:
        return RedirectResponse("/login?err=forbidden", status_code=303)
    return None

# ----------------------------
# Routes: Root & HMRC OAuth
# ----------------------------
@app.get("/")
def root():
    return {
        "ok": True,
        "connect": "/connect",
        "portal": "/portal",
        "dashboard": "/dashboard",
        "login": "/login",
        "register_agent": "/register/agent",
        "register_taxpayer": "/register/taxpayer",
        "endpoints": [
            "/api/obligations?vrn=VRN",
            "/api/returns (POST)",
            "/api/returns/view?vrn=VRN&periodKey=18A1",
            "/api/liabilities?vrn=VRN&from_=YYYY-MM-DD&to=YYYY-MM-DD",
            "/api/payments?vrn=VRN&from_=YYYY-MM-DD&to=YYYY-MM-DD",
            "/api/receipts",
            "/api/excel/preview (POST)",
        ],
    }

@app.get("/connect")
def connect():
    state = str(uuid.uuid4())
    STORE["state"] = state
    return RedirectResponse(auth_url(state))

@app.get("/oauth/hmrc/callback")
async def oauth_callback(code: str, state: str):
    if state != STORE.get("state"):
        raise HTTPException(400, "state mismatch")

    tokens = await token_request(
        {
            "grant_type": "authorization_code",
            "client_id": HMRC_CLIENT_ID,
            "client_secret": HMRC_CLIENT_SECRET,
            "redirect_uri": HMRC_REDIRECT_URI,
            "code": code,
        }
    )
    tokens["obtained_at"] = time.time()
    STORE["tokens"] = tokens
    save_tokens(tokens)
    return PlainTextResponse(
        f"Connected. Access token received. Expires in {tokens.get('expires_in')} seconds."
    )

# ----------------------------
# JSON API: VAT endpoints
# ----------------------------
@app.get("/api/obligations")
async def obligations(request: Request, vrn: str, status: str = "O", scenario: Optional[str] = None):
    tok = await access_token()
    url = f"{BASE_URL}/organisations/vat/{vrn}/obligations"
    headers = {**hmrc_headers(request), "Authorization": f"Bearer {tok}"}
    if scenario:
        headers["Gov-Test-Scenario"] = scenario

    async with httpx.AsyncClient() as client:
        r = await client.get(url, params={"status": status}, headers=headers)
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)
    return r.json() if r.text else {}

@app.post("/api/returns")
async def submit_return(request: Request, vrn: str, payload: dict):
    tok = await access_token()
    url = f"{BASE_URL}/organisations/vat/{vrn}/returns"
    headers = {
        **hmrc_headers(request),
        "Authorization": f"Bearer {tok}",
        "Content-Type": "application/json",
    }
    async with httpx.AsyncClient() as client:
        r = await client.post(url, json=payload, headers=headers)
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)

    resp = r.json()
    period_key = payload.get("periodKey") if isinstance(payload, dict) else None
    append_receipt(vrn, period_key, resp)
    return resp

@app.get("/api/returns/view")
async def view_return(request: Request, vrn: str, periodKey: str):
    tok = await access_token()
    url = f"{BASE_URL}/organisations/vat/{vrn}/returns/{periodKey}"
    async with httpx.AsyncClient() as client:
        r = await client.get(url, headers={**hmrc_headers(request), "Authorization": f"Bearer {tok}"})
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)
    return r.json()

@app.get("/api/liabilities")
async def liabilities(request: Request, vrn: str, from_: str, to: str, scenario: Optional[str] = None):
    tok = await access_token()
    url = f"{BASE_URL}/organisations/vat/{vrn}/liabilities"
    headers = {**hmrc_headers(request), "Authorization": f"Bearer {tok}"}
    if scenario:
        headers["Gov-Test-Scenario"] = scenario
    async with httpx.AsyncClient() as client:
        r = await client.get(url, params={"from": from_, "to": to}, headers=headers)
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)
    return r.json()

@app.get("/api/payments")
async def payments(request: Request, vrn: str, from_: str, to: str, scenario: Optional[str] = None):
    tok = await access_token()
    url = f"{BASE_URL}/organisations/vat/{vrn}/payments"
    headers = {**hmrc_headers(request), "Authorization": f"Bearer {tok}"}
    if scenario:
        headers["Gov-Test-Scenario"] = scenario
    async with httpx.AsyncClient() as client:
        r = await client.get(url, params={"from": from_, "to": to}, headers=headers)
    if r.status_code >= 400:
        raise HTTPException(r.status_code, r.text)
    return r.json()

@app.get("/api/receipts")
def receipts():
    return load_json(RECEIPTS_FILE, [])

# ----------------------------
# Excel preview API
# ----------------------------
@app.post("/api/excel/preview")
async def excel_preview(
    file: UploadFile = File(...),
    box1: str = Form(...),
    box2: str = Form(...),
    box4: str = Form(...),
    box6: str = Form(...),
    box7: str = Form(...),
    box8: str = Form(...),
    box9: str = Form(...),
):
    """
    Reads the uploaded .xlsx and returns values for VAT boxes based on cell addresses.
    e.g. box1="B2" means read cell B2 on the first sheet.
    """
    content = await file.read()
    tmp = pathlib.Path("_upload.xlsx")
    tmp.write_bytes(content)
    try:
        wb = load_workbook(tmp, data_only=True)
        ws = wb.active

        def f(c):
            v = ws[c].value
            try:
                return float(v)
            except Exception:
                return 0.0

        vatDueSales = f(box1)
        vatDueAcquisitions = f(box2)
        vatReclaimedCurrPeriod = f(box4)
        totalValueSalesExVAT = f(box6)
        totalValuePurchasesExVAT = f(box7)
        totalValueGoodsSuppliedExVAT = f(box8)
        totalAcquisitionsExVAT = f(box9)

        totalVatDue = vatDueSales + vatDueAcquisitions
        netVatDue = totalVatDue - vatReclaimedCurrPeriod

        return {
            "vatDueSales": round(vatDueSales, 2),
            "vatDueAcquisitions": round(vatDueAcquisitions, 2),
            "totalVatDue": round(totalVatDue, 2),
            "vatReclaimedCurrPeriod": round(vatReclaimedCurrPeriod, 2),
            "netVatDue": round(netVatDue, 2),
            "totalValueSalesExVAT": int(totalValueSalesExVAT),
            "totalValuePurchasesExVAT": int(totalValuePurchasesExVAT),
            "totalValueGoodsSuppliedExVAT": int(totalValueGoodsSuppliedExVAT),
            "totalAcquisitionsExVAT": int(totalAcquisitionsExVAT),
        }
    finally:
        if tmp.exists():
            tmp.unlink(missing_ok=True)

# ----------------------------
# Auth UI (Register / Login / Logout)
# ----------------------------
HEADER_HTML = """
<script src="https://cdn.tailwindcss.com"></script>
"""

def layout(title: str, body: str) -> HTMLResponse:
    html = f"""<!doctype html><html><head>
      <meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
      <title>{title}</title>
      {HEADER_HTML}
    </head>
    <body class="min-h-screen bg-slate-50 text-slate-900">
      <div class="max-w-4xl mx-auto px-4 py-8">
        {body}
      </div>
    </body></html>"""
    return HTMLResponse(html)

def auth_card(title: str, content: str) -> str:
    return f"""
    <div class="bg-white border border-slate-200 rounded-xl shadow-sm p-6">
      <h1 class="text-2xl font-semibold mb-4">{title}</h1>
      {content}
    </div>
    """

@app.get("/portal", response_class=HTMLResponse)
def portal(request: Request):
    u = current_user(request)
    body = f"""
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-3xl font-semibold">My VAT Filer</h1>
      <div class="text-sm">
        {"Signed in as <b>"+u["email"]+"</b> ("+u["role"]+") · <a class='text-blue-700' href='/logout'>Logout</a>" if u else "<a class='text-blue-700' href='/login'>Login</a>"}
      </div>
    </div>

    <div class="grid md:grid-cols-2 gap-4">
      <a href="/register/agent" class="block p-5 rounded-xl border bg-white hover:bg-slate-50">
        <div class="text-lg font-medium">Register as Tax Agent</div>
        <div class="text-slate-600 text-sm">Accountants/Agents managing multiple businesses.</div>
      </a>
      <a href="/register/taxpayer" class="block p-5 rounded-xl border bg-white hover:bg-slate-50">
        <div class="text-lg font-medium">Register as Taxpayer</div>
        <div class="text-slate-600 text-sm">Business filing your own VAT returns.</div>
      </a>
      <a href="/login" class="block p-5 rounded-xl border bg-white hover:bg-slate-50">
        <div class="text-lg font-medium">Login</div>
        <div class="text-slate-600 text-sm">Already have an account? Sign in.</div>
      </a>
      <a href="/dashboard" class="block p-5 rounded-xl border bg-white hover:bg-slate-50">
        <div class="text-lg font-medium">Dashboard</div>
        <div class="text-slate-600 text-sm">Pick obligations and prepare/submit a return.</div>
      </a>
    </div>
    """
    return layout("Portal – My VAT Filer", body)

def _register_form(role_label: str, role_value: str, err: str = "") -> str:
    return auth_card(
        f"{role_label} Account Signup",
        f"""
<form method="post" class="space-y-3">
  {"<div class='text-red-700 bg-red-50 border border-red-200 rounded p-2 text-sm'>"+err+"</div>" if err else ""}
  <div>
    <label class="block text-sm font-medium mb-1">Email address</label>
    <input name="email" type="email" required class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500" placeholder="you@example.com">
  </div>
  <div class="grid md:grid-cols-2 gap-3">
    <div>
      <label class="block text-sm font-medium mb-1">Password</label>
      <input name="password" type="password" minlength="8" required class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
      <div class="text-xs text-slate-500 mt-1">Please don't use  &lt; &gt; , &amp; \" '  in your password.</div>
    </div>
    <div>
      <label class="block text-sm font-medium mb-1">Confirm Password</label>
      <input name="confirm" type="password" minlength="8" required class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
    </div>
  </div>
  <div class="grid md:grid-cols-2 gap-3">
    <div>
      <label class="block text-sm font-medium mb-1">Contact Name</label>
      <input name="contact_name" class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
    </div>
    <div>
      <label class="block text-sm font-medium mb-1">Business Name</label>
      <input name="business_name" class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
    </div>
  </div>
  <div>
    <label class="block text-sm font-medium mb-1">Phone Number</label>
    <input name="phone" class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
  </div>

  <input type="hidden" name="role" value="{role_value}">
  <div class="pt-2 flex items-center gap-2">
    <a href="/portal" class="rounded border px-4 py-2 hover:bg-slate-50">Cancel</a>
    <button class="rounded bg-blue-600 text-white px-5 py-2 hover:bg-blue-700">Register</button>
  </div>
</form>
"""
    )

@app.get("/register/agent", response_class=HTMLResponse)
def register_agent():
    return layout("Register – Tax Agent", _register_form("Tax Agent", "agent"))

@app.post("/register/agent", response_class=HTMLResponse)
async def register_agent_post(request: Request):
    form = await request.form()
    try:
        if form.get("password") != form.get("confirm"):
            raise ValueError("Passwords do not match.")
        create_user(
            email=str(form.get("email", "")),
            password=str(form.get("password", "")),
            role="agent",
            contact_name=str(form.get("contact_name", "")),
            business_name=str(form.get("business_name", "")),
            phone=str(form.get("phone", "")),
        )
        # Auto-login
        request.session["user"] = {"email": str(form.get("email", "")).lower()}
        return RedirectResponse("/dashboard", status_code=303)
    except Exception as e:
        return layout("Register – Tax Agent", _register_form("Tax Agent", "agent", str(e)))

@app.get("/register/taxpayer", response_class=HTMLResponse)
def register_taxpayer():
    return layout("Register – Taxpayer", _register_form("Taxpayer", "taxpayer"))

@app.post("/register/taxpayer", response_class=HTMLResponse)
async def register_taxpayer_post(request: Request):
    form = await request.form()
    try:
        if form.get("password") != form.get("confirm"):
            raise ValueError("Passwords do not match.")
        create_user(
            email=str(form.get("email", "")),
            password=str(form.get("password", "")),
            role="taxpayer",
            contact_name=str(form.get("contact_name", "")),
            business_name=str(form.get("business_name", "")),
            phone=str(form.get("phone", "")),
        )
        # Auto-login
        request.session["user"] = {"email": str(form.get("email", "")).lower()}
        return RedirectResponse("/dashboard", status_code=303)
    except Exception as e:
        return layout("Register – Taxpayer", _register_form("Taxpayer", "taxpayer", str(e)))

def _login_form(err: str = "") -> str:
    return auth_card(
        "Sign in",
        f"""
<form method="post" class="space-y-3">
  {"<div class='text-red-700 bg-red-50 border border-red-200 rounded p-2 text-sm'>"+err+"</div>" if err else ""}
  <div>
    <label class="block text-sm font-medium mb-1">Email</label>
    <input name="email" type="email" required class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
  </div>
  <div>
    <label class="block text-sm font-medium mb-1">Password</label>
    <input name="password" type="password" required class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
  </div>
  <div class="pt-2 flex items-center gap-2">
    <a href="/portal" class="rounded border px-4 py-2 hover:bg-slate-50">Cancel</a>
    <button class="rounded bg-blue-600 text-white px-5 py-2 hover:bg-blue-700">Login</button>
  </div>
</form>
<div class="mt-4 text-sm">
  <a class="text-blue-700" href="/register/taxpayer">Create a Taxpayer account</a> ·
  <a class="text-blue-700" href="/register/agent">Create a Tax Agent account</a>
</div>
"""
    )

@app.get("/login", response_class=HTMLResponse)
def login_get(request: Request, err: Optional[str] = None):
    return layout("Login – My VAT Filer", _login_form("You don’t have permission for that page." if err == "forbidden" else ""))

@app.post("/login")
async def login_post(request: Request):
    form = await request.form()
    email = str(form.get("email", "")).lower().strip()
    password = str(form.get("password", ""))
    u = get_user_by_email(email)
    if not u or not verify_password(password, u):
        return layout("Login – My VAT Filer", _login_form("Invalid email or password."))
    request.session["user"] = {"email": email}
    return RedirectResponse("/dashboard", status_code=303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/portal", status_code=303)

# ----------------------------
# Dashboard / Prepare / UI (protected)
# ----------------------------
@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    redir = require_login(request)  # any role
    if redir: return redir
    u = current_user(request)
    html = f"""
<!doctype html><html><head>
  <meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Dashboard – My VAT Filer</title>
  {HEADER_HTML}
</head><body class="bg-slate-50">
  <div class="max-w-4xl mx-auto px-4 py-8">
    <div class="flex items-center justify-between mb-6">
      <h1 class="text-2xl font-semibold">Dashboard</h1>
      <div class="text-sm text-slate-600">
        Signed in as <b>{u['email']}</b> ({u['role']}) · <a class="text-blue-700" href="/logout">Logout</a>
      </div>
    </div>

    <div class="bg-white border border-slate-200 rounded-xl p-5 shadow">
      <div class="grid md:grid-cols-4 gap-3 items-end">
        <div class="md:col-span-2">
          <label class="block text-sm font-medium mb-1">VRN</label>
          <input id="vrn" class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500" placeholder="e.g. 458814905"/>
        </div>
        <div>
          <label class="block text-sm font-medium mb-1">Sandbox scenario</label>
          <select id="scenario" class="w-full rounded border-slate-300 focus:border-blue-500 focus:ring-blue-500">
            <option value="">(default)</option>
            <option>QUARTERLY_NONE_MET</option>
            <option>QUARTERLY_ONE_MET</option>
            <option>MULTIPLE_OBLIGATIONS</option>
          </select>
        </div>
        <div>
          <button id="load" class="w-full rounded bg-blue-600 text-white py-2 hover:bg-blue-700">Load</button>
        </div>
      </div>

      <div id="list" class="mt-5"></div>
    </div>
  </div>

<script>
const $ = (id)=>document.getElementById(id);
$('load').onclick = async ()=>{
  const vrn = $('vrn').value.trim();
  const sc = $('scenario').value.trim();
  if(!vrn){ alert('Enter VRN'); return; }
  const url = sc ? `/api/obligations?vrn=${vrn}&scenario=${encodeURIComponent(sc)}`
                 : `/api/obligations?vrn=${vrn}`;
  const r = await fetch(url);
  const data = await r.json();
  const obs = data.obligations || [];
  const list = $('list');
  list.innerHTML = '';
  if(!obs.length){ list.textContent = 'No open obligations.'; return; }
  obs.forEach(o=>{
    const a = document.createElement('a');
    a.className='block border rounded p-3 mb-2 bg-white hover:bg-slate-50';
    a.href=`/prepare?vrn=${vrn}&periodKey=${encodeURIComponent(o.periodKey)}`;
    a.textContent = `${o.periodKey} · ${o.start} → ${o.end} · due ${o.due}  —  Prepare from Excel`;
    list.appendChild(a);
  });
};
</script>
</body></html>
"""
    return HTMLResponse(html)

@app.get("/prepare", response_class=HTMLResponse)
def prepare(request: Request, vrn: str, periodKey: str):
    redir = require_login(request)  # any role
    if redir: return redir
    html = f"""
<!doctype html><html><head>
  <meta charset="utf-8"/><meta name="viewport" content="width=device-width, initial-scale=1"/>
  <title>Prepare from Excel – My VAT Filer</title>
  {HEADER_HTML}
</head><body class="bg-slate-50">
  <div class="max-w-3xl mx-auto px-4 py-8">
    <a class="text-sm text-blue-700" href="/dashboard">← Back to dashboard</a>
    <h1 class="text-2xl font-semibold mt-2">Prepare return</h1>
    <p class="text-slate-600">VRN <b>{vrn}</b>, period <b>{periodKey}</b></p>

    <form id="f" class="bg-white border border-slate-200 rounded-xl p-5 mt-4 space-y-3" enctype="multipart/form-data">
      <div>
        <label class="block text-sm font-medium mb-1">Excel file (.xlsx)</label>
        <input type="file" name="file" accept=".xlsx" required/>
      </div>

      <p class="text-sm text-slate-600">Enter cell addresses (first sheet). Examples: B2, C10, F7.</p>

      <div class="grid md:grid-cols-2 gap-3">
        <label>Box 1 cell <input name="box1" class="w-full rounded border-slate-300" value="B2"/></label>
        <label>Box 2 cell <input name="box2" class="w-full rounded border-slate-300" value="B3"/></label>
        <label>Box 4 cell <input name="box4" class="w-full rounded border-slate-300" value="B4"/></label>
        <label>Box 6 cell <input name="box6" class="w-full rounded border-slate-300" value="B5"/></label>
        <label>Box 7 cell <input name="box7" class="w-full rounded border-slate-300" value="B6"/></label>
        <label>Box 8 cell <input name="box8" class="w-full rounded border-slate-300" value="B7"/></label>
        <label>Box 9 cell <input name="box9" class="w-full rounded border-slate-300" value="B8"/></label>
      </div>

      <button class="rounded bg-blue-600 text-white py-2 px-4 hover:bg-blue-700">Preview</button>
    </form>

    <pre id="out" class="mt-4 text-sm bg-white border border-slate-200 rounded p-3"></pre>

    <button id="use" class="hidden mt-3 rounded bg-green-600 text-white py-2 px-4 hover:bg-green-700">Use these values → Go to /ui</button>
  </div>

<script>
const params = new URLSearchParams(location.search);
const vrn = params.get('vrn');
const periodKey = params.get('periodKey');
const out = document.getElementById('out');
const use = document.getElementById('use');

document.getElementById('f').onsubmit = async (e)=>{
  e.preventDefault();
  const fd = new FormData(e.target);
  const r = await fetch('/api/excel/preview', { method:'POST', body: fd });
  const data = await r.json();
  out.textContent = JSON.stringify(data, null, 2);
  if(r.ok) {
    use.classList.remove('hidden');
    data.periodKey = periodKey;
    localStorage.setItem('prefill', JSON.stringify(data));
  }
};
use.onclick = ()=>{ window.location = '/ui?vrn='+encodeURIComponent(vrn); };
</script>
</body></html>
"""
    return HTMLResponse(html)

@app.get("/ui", response_class=HTMLResponse)
def ui(request: Request):
    redir = require_login(request)  # any role
    if redir: return redir
    # This is the same UI you had previously (trimmed only where not essential)
    return HTMLResponse("""
<!doctype html>
<html lang="en" class="h-full">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>My VAT Filer — Sandbox</title>
  <script src="https://cdn.tailwindcss.com"></script>
  <script>tailwind.config = { theme: { extend: { colors: { brand:'#2563eb' }}}}</script>
</head>
<body class="h-full bg-slate-50 text-slate-900">
  <div class="max-w-6xl mx-auto px-4 py-8">
    <header class="mb-8">
      <h1 class="text-3xl font-semibold tracking-tight">My VAT Filer <span class="text-slate-400">(Sandbox)</span></h1>
      <p class="text-slate-600 mt-1">Connect, pick an obligation period, complete Boxes 1–9, submit, and view the receipt.</p>
      <p class="text-sm mt-1"><a class="text-blue-700" href="/dashboard">Go to dashboard</a></p>
    </header>

    <div id="toast" class="hidden fixed top-4 right-4 z-50 min-w-[280px] rounded-md border border-slate-200 bg-white shadow-lg p-3"></div>

    <div class="grid md:grid-cols-3 gap-6">
      <section class="md:col-span-1 rounded-xl bg-white border border-slate-200 shadow-sm p-5">
        <h2 class="font-medium text-slate-800 mb-3">1) Connect to HMRC</h2>
        <p class="text-sm text-slate-600 mb-4">Use your sandbox test organisation.</p>
        <button id="btnConnect" class="inline-flex items-center justify-center rounded-lg bg-brand px-4 py-2 text-white hover:bg-blue-600 active:bg-blue-700 transition">Connect</button>
        <div id="status" class="mt-3 text-sm text-slate-600">Status: <span class="font-medium">Unknown</span></div>
      </section>

      <section class="md:col-span-2 rounded-xl bg-white border border-slate-200 shadow-sm p-5">
        <h2 class="font-medium text-slate-800 mb-4">2) Pick an obligation</h2>

        <div class="grid md:grid-cols-4 gap-3 items-end">
          <div class="md:col-span-2">
            <label class="block text-sm font-medium text-slate-700 mb-1">VRN</label>
            <input id="vrn" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" placeholder="e.g. 458814905" />
          </div>

          <div>
            <label class="block text-sm font-medium text-slate-700 mb-1">Scenario (sandbox)</label>
            <select id="scenario" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand">
              <option value="">(default)</option>
              <option>QUARTERLY_NONE_MET</option>
              <option>QUARTERLY_ONE_MET</option>
              <option>MULTIPLE_OBLIGATIONS</option>
            </select>
          </div>

          <div class="flex gap-2">
            <button id="btnLoad" class="flex-1 rounded-lg border border-slate-300 hover:bg-slate-50 px-4 py-2 transition">Load obligations</button>
          </div>
        </div>

        <div class="grid md:grid-cols-3 gap-3 mt-4">
          <div>
            <label class="block text-sm font-medium text-slate-700 mb-1">Obligation (periodKey)</label>
            <select id="periodSelect" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand">
              <option value="">— none loaded —</option>
            </select>
          </div>
          <div>
            <label class="block text-sm font-medium text-slate-700 mb-1">Auto-filled periodKey</label>
            <input id="periodKey" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand bg-slate-50" readonly />
          </div>
          <div class="text-sm text-slate-600 flex items-end" id="obligMeta"></div>
        </div>
      </section>
    </div>

    <section class="rounded-xl bg-white border border-slate-200 shadow-sm p-5 mt-6">
      <h2 class="font-medium text-slate-800 mb-2">3) Fill & Submit VAT return</h2>
      <p class="text-xs text-slate-500 mb-4">Boxes 2, 8 and 9 relate to EU movements. In most cases set them to 0.</p>

      <div class="grid grid-cols-1 gap-y-5">
        <div><label class="block text-sm font-medium mb-1">Box 1 — VAT due on sales</label><input id="vatDueSales" value="100.00" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 2 — VAT due on acquisitions (EU)</label><input id="vatDueAcquisitions" value="0.00" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 3 — Total VAT due (auto)</label><input id="totalVatDue" value="100.00" class="w-full rounded-lg border-slate-300" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 4 — VAT reclaimed on inputs</label><input id="vatReclaimedCurrPeriod" value="0.00" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 5 — Net VAT (auto)</label><input id="netVatDue" value="100.00" class="w-full rounded-lg border-slate-300" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 6 — Total value of sales (ex VAT)</label><input id="totalValueSalesExVAT" value="500" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 7 — Total value of purchases (ex VAT)</label><input id="totalValuePurchasesExVAT" value="0" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 8 — Supplies to EU (ex VAT)</label><input id="totalValueGoodsSuppliedExVAT" value="0" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Box 9 — Acquisitions from EU (ex VAT)</label><input id="totalAcquisitionsExVAT" value="0" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand" /></div>
        <div><label class="block text-sm font-medium mb-1">Declaration</label><select id="finalised" class="w-full rounded-lg border-slate-300 focus:border-brand focus:ring-brand"><option>true</option><option>false</option></select></div>
      </div>

      <div class="mt-5">
        <button id="btnSubmit" class="rounded-lg bg-brand text-white px-5 py-2 hover:bg-blue-600 active:bg-blue-700 transition">Submit return</button>
      </div>
    </section>

    <section class="rounded-xl bg-white border border-slate-200 shadow-sm p-5 mt-6">
      <div class="flex items-center justify-between">
        <h2 class="font-medium text-slate-800">Result</h2>
        <div class="flex gap-2">
          <button id="btnCopy" class="rounded-lg border border-slate-300 px-3 py-1.5 hover:bg-slate-50">Copy</button>
          <button id="btnClear" class="rounded-lg border border-slate-300 px-3 py-1.5 hover:bg-slate-50">Clear</button>
        </div>
      </div>
      <pre id="out" class="mt-4 text-sm whitespace-pre-wrap bg-slate-50 border border-slate-200 rounded-lg p-3"></pre>
    </section>
  </div>

<script>
const $ = (id) => document.getElementById(id);
const out = $('out');
const toast = document.getElementById('toast');

function notify(msg, type='info'){
  toast.className = "fixed top-4 right-4 z-50 min-w-[280px] rounded-md border p-3 shadow-lg " +
    (type==='error' ? "bg-red-50 border-red-200 text-red-700"
     : type==='success' ? "bg-green-50 border-green-200 text-green-700"
     : "bg-white border-slate-200 text-slate-800");
  toast.textContent = msg;
  toast.classList.remove('hidden');
  setTimeout(()=> toast.classList.add('hidden'), 3000);
}
function pretty(objOrText){
  try{ return JSON.stringify(typeof objOrText==='string' ? JSON.parse(objOrText) : objOrText, null, 2); }
  catch{ return String(objOrText); }
}
async function getJSON(url){
  const r = await fetch(url);
  if(!r.ok) throw new Error(await r.text());
  return r.json();
}
function recalcTotals(){
  const vds = parseFloat($('vatDueSales').value || 0);
  const vda = parseFloat($('vatDueAcquisitions').value || 0);
  const vrc = parseFloat($('vatReclaimedCurrPeriod').value || 0);
  $('totalVatDue').value = (vds + vda).toFixed(2);
  $('netVatDue').value   = (vds + vda - vrc).toFixed(2);
}
['vatDueSales','vatDueAcquisitions','vatReclaimedCurrPeriod'].forEach(id=>{
  $(id).addEventListener('input', recalcTotals);
});

// prefill from /prepare (values are in localStorage.prefill)
try {
  const pf = JSON.parse(localStorage.getItem('prefill') || '{}');
  if (pf && Object.keys(pf).length) {
    if (pf.periodKey) $('periodKey').value = pf.periodKey;
    if ('vatDueSales' in pf) $('vatDueSales').value = Number(pf.vatDueSales).toFixed(2);
    if ('vatDueAcquisitions' in pf) $('vatDueAcquisitions').value = Number(pf.vatDueAcquisitions).toFixed(2);
    if ('vatReclaimedCurrPeriod' in pf) $('vatReclaimedCurrPeriod').value = Number(pf.vatReclaimedCurrPeriod).toFixed(2);
    if ('totalValueSalesExVAT' in pf) $('totalValueSalesExVAT').value = pf.totalValueSalesExVAT;
    if ('totalValuePurchasesExVAT' in pf) $('totalValuePurchasesExVAT').value = pf.totalValuePurchasesExVAT;
    if ('totalValueGoodsSuppliedExVAT' in pf) $('totalValueGoodsSuppliedExVAT').value = pf.totalValueGoodsSuppliedExVAT;
    if ('totalAcquisitionsExVAT' in pf) $('totalAcquisitionsExVAT').value = pf.totalAcquisitionsExVAT;
    recalcTotals();
    notify('Values loaded from Excel preview','success');
    localStorage.removeItem('prefill');
  }
} catch(e) { /* ignore */ }

// Connect
$('btnConnect').onclick = () => { window.location = '/connect'; };

// Load obligations
$('btnLoad').onclick = async ()=>{
  try{
    const vrn = $('vrn').value.trim();
    const scenario = $('scenario').value;
    if(!vrn){ notify('Enter a VRN first','error'); return; }

    const url = scenario ? `/api/obligations?vrn=${vrn}&scenario=${encodeURIComponent(scenario)}`
                         : `/api/obligations?vrn=${vrn}`;
    const data = await getJSON(url);

    const select = $('periodSelect');
    select.innerHTML = '';
    const obs = data.obligations || [];
    if(!obs.length){
      select.innerHTML = '<option value="">— no open obligations —</option>';
      $('periodKey').value = '';
      $('obligMeta').textContent = '';
      notify('No open obligations returned','info');
      return;
    }

    obs.forEach(o=>{
      const opt = document.createElement('option');
      opt.value = o.periodKey;
      opt.textContent = `${o.periodKey} · ${o.start} → ${o.end} · due ${o.due}`;
      opt.dataset.meta = JSON.stringify(o);
      select.appendChild(opt);
    });

    select.onchange = ()=>{
      const meta = JSON.parse(select.selectedOptions[0].dataset.meta);
      $('periodKey').value = meta.periodKey;
      $('obligMeta').textContent = `Selected: ${meta.start} → ${meta.end}, due ${meta.due}`;
    };
    select.dispatchEvent(new Event('change'));
    notify('Obligations loaded','success');
  }catch(e){
    out.textContent = pretty(e.message);
    notify('Failed to load obligations','error');
  }
};

// Submit return
$('btnSubmit').onclick = async ()=>{
  try{
    const vrn = $('vrn').value.trim();
    const periodKey = $('periodKey').value.trim();
    if(!vrn || !periodKey){ notify('VRN and periodKey required','error'); return; }

    const pre = await fetch(`/api/returns/view?vrn=${vrn}&periodKey=${encodeURIComponent(periodKey)}`);
    if (pre.ok) {
      const already = await pre.json();
      out.textContent = 'Already submitted. HMRC shows:\\n' + pretty(already);
      notify('Already submitted for this period','info');
      return;
    }

    recalcTotals();
    const body = {
      periodKey,
      vatDueSales: $('vatDueSales').value.trim(),
      vatDueAcquisitions: $('vatDueAcquisitions').value.trim(),
      totalVatDue: $('totalVatDue').value.trim(),
      vatReclaimedCurrPeriod: $('vatReclaimedCurrPeriod').value.trim(),
      netVatDue: $('netVatDue').value.trim(),
      totalValueSalesExVAT: parseInt($('totalValueSalesExVAT').value,10) || 0,
      totalValuePurchasesExVAT: parseInt($('totalValuePurchasesExVAT').value,10) || 0,
      totalValueGoodsSuppliedExVAT: parseInt($('totalValueGoodsSuppliedExVAT').value,10) || 0,
      totalAcquisitionsExVAT: parseInt($('totalAcquisitionsExVAT').value,10) || 0,
      finalised: $('finalised').value === 'true'
    };

    const r = await fetch(`/api/returns?vrn=${vrn}`, {
      method:'POST',
      headers:{'Content-Type':'application/json'},
      body: JSON.stringify(body)
    });
    const txt = await r.text();
    out.textContent = pretty(txt);
    if(r.ok){ notify('Submitted successfully','success'); }
    else{ notify('Submission returned an error','error'); }
  }catch(e){
    out.textContent = pretty(e.message);
    notify('Submit failed','error');
  }
};

$('btnCopy').onclick = async ()=>{ await navigator.clipboard.writeText(out.textContent || ''); notify('Copied to clipboard','success'); };
$('btnClear').onclick = ()=>{ out.textContent = ''; };
</script>
</body>
</html>
""")

# ----------------------------
# Done
# ----------------------------
